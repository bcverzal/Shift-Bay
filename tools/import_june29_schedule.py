import json
import re
import shutil
import time
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "restaurant-scheduler-data.json"
BACKUP_DIR = ROOT / "data" / "backups"
SOURCE_XLSX = Path(r"C:\Users\bcver\Downloads\june 29.xlsx")
WEEK_DATES = {
    1: "2026-06-23",
    2: "2026-06-24",
    3: "2026-06-25",
    4: "2026-06-26",
    5: "2026-06-27",
    6: "2026-06-28",
    7: "2026-06-29",
}
WEEK_START = "2026-06-23"
WEEK_END = "2026-06-29"

ROLE_ALIASES = {
    "SERVER": "Server",
    "SERVE": "Server",
    "HOST": "Host",
    "BUS": "Busser",
    "BUSSER": "Busser",
    "BAR": "Bartender",
    "BARTENDER": "Bartender",
    "EXPO": "Expo",
    "BQT": "Banquet Server",
    "BANQUET": "Banquet Server",
    "BUFFET": "BOH Block",
}

NAME_ALIASES = {
    "AJ": "Todd Cruz",
    "Brian": "Brian Sass",
    "Chole": "Chloe",
    "Gaby": "Gabriella Diaz",
    "Jack": "Jackson Lindenberg",
    "Lily": "Lilian Hatzung",
    "Lito": "Lito Ortega",
    "Nellie": "Antonella Demarco",
    "Patty": "Patricia Petretti",
    "Reese": "Teresa Fechter",
    "Sam": "Samantha Casper",
}

CROSS_REFERENCE_WORDS = {"BAR", "HOST", "BUS", "BUSSER", "SERVE", "SERVER", "EXPO"}


def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def uid(prefix):
    return f"{prefix}_{int(time.time() * 1000):x}_{str(time.perf_counter_ns())[-6:]}"


def normalize_key(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def full_name(employee):
    return f"{employee.get('firstName', '')} {employee.get('lastName', '')}".strip()


def build_employee_lookup(employees):
    lookup = {}
    for employee in employees:
        names = {
            employee.get("firstName", ""),
            employee.get("nickname", ""),
            full_name(employee),
        }
        for name in names:
            key = normalize_key(name)
            if key:
                lookup.setdefault(key, employee)
    return lookup


def find_employee(data, source_name, created):
    target_name = NAME_ALIASES.get(source_name, source_name)
    lookup = build_employee_lookup(data["employees"])
    employee = lookup.get(normalize_key(target_name))
    if employee:
        return employee
    return None


def role_by_name(data):
    return {role["name"].lower(): role for role in data["roles"]}


def ensure_role_training(employee, role):
    if role["department"] == "Exec":
        return
    employee.setdefault("roleTraining", [])
    if role["id"] not in employee["roleTraining"]:
        employee["roleTraining"].append(role["id"])
    employee.setdefault("departments", [])
    if role["department"] not in employee["departments"]:
        employee["departments"].append(role["department"])


def parse_time_token(raw, date_key, default_role_name):
    text = str(raw).strip().upper().replace(" ", "")
    text = text.replace("TRAIN", " TRAIN").replace("TESTDAY", " TEST DAY")
    is_training = "TRAIN" in text or "TEST DAY" in text
    notes = []
    if "TRAIN" in text:
        notes.append("Training")
    if "TEST DAY" in text:
        notes.append("Test day")
    if "CHAIRS" in text:
        notes.append("Chairs")
    clean = re.sub(r"TRAIN|TEST\s*DAY|CHAIRS", "", text).strip()
    role_name = default_role_name
    for marker, mapped in ROLE_ALIASES.items():
        if marker in clean:
            role_name = mapped
            clean = clean.replace(marker, "")
            break
    is_closer = "CL" in clean
    clean = clean.replace("CL", "")
    is_flex = "?" in clean
    clean = clean.replace("?", "")
    if clean == "":
        return None
    if "-" in clean:
        start_raw, end_raw = clean.split("-", 1)
    else:
        start_raw, end_raw = clean, ""
    start = normalize_time(start_raw, prefer_pm=default_start_pm(role_name, start_raw, date_key))
    if not start:
        return None
    if end_raw:
        end = normalize_time(end_raw, prefer_pm=end_should_be_pm(start, end_raw))
        until_volume = False
    elif is_flex:
        end = "Until Volume"
        until_volume = True
    else:
        end = default_end_for(role_name, start, date_key, is_closer)
        until_volume = False
    return {
        "roleName": role_name,
        "start": start,
        "end": end,
        "untilVolume": until_volume,
        "isCloser": is_closer,
        "isFlexDouble": is_flex,
        "isTraining": is_training,
        "notes": "; ".join(notes),
    }


def default_start_pm(role_name, raw, date_key):
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return False
    num = int(digits)
    if role_name in {"Bartender", "Expo"}:
        return True if num in {3, 4, 430, 5, 530} else False
    if role_name in {"Host", "Busser"} and num in {3, 4, 430, 5, 530}:
        return True
    if role_name == "Server" and num in {3, 4, 430, 5, 530}:
        return True
    if role_name == "Banquet Server" and num in {3, 4, 430, 5, 530}:
        return True
    return False


def normalize_time(raw, prefer_pm=False):
    value = str(raw or "").strip().upper().replace(":", "")
    match = re.match(r"^(\d{1,4})(AM|PM)?$", value)
    if not match:
        return ""
    digits, suffix = match.groups()
    if len(digits) <= 2:
        hour = int(digits)
        minute = 0
    else:
        hour = int(digits[:-2])
        minute = int(digits[-2:])
    if suffix:
        ampm = suffix
    else:
        ampm = "PM" if prefer_pm else "AM"
    if hour == 0:
        hour = 12
    return f"{hour}:{minute:02d} {ampm}"


def minutes(time_text):
    match = re.match(r"^(\d{1,2}):(\d{2})\s*(AM|PM)$", time_text)
    if not match:
        return 0
    hour, minute, suffix = match.groups()
    hour = int(hour) % 12
    if suffix == "PM":
        hour += 12
    return hour * 60 + int(minute)


def end_should_be_pm(start, raw_end):
    end_digits = int(re.sub(r"\D", "", str(raw_end)) or 0)
    start_minutes = minutes(start)
    if start_minutes >= 12 * 60:
        return True
    return end_digits in {1, 130, 2, 230, 3, 330, 4, 430, 5, 530, 6, 630, 7, 730, 8, 830}


def default_end_for(role_name, start, date_key, is_closer=False):
    start_minutes = minutes(start)
    day = int(date_key[-2:])
    weekend = day in {26, 27, 28}
    if role_name == "BOH Block":
        return "4:00 PM"
    if role_name == "Expo":
        return "8:00 PM"
    if role_name == "Bartender":
        if start_minutes < 12 * 60:
            return "2:00 PM"
        return "10:00 PM" if weekend else "8:00 PM"
    if role_name == "Host":
        if start_minutes < 12 * 60:
            return "8:00 PM" if start_minutes >= 9 * 60 else "3:00 PM"
        return "8:30 PM"
    if role_name == "Busser":
        if start_minutes < 12 * 60:
            return "2:00 PM"
        return "8:00 PM" if not weekend else "10:00 PM"
    if role_name == "Banquet Server":
        return "9:00 PM" if start_minutes >= 12 * 60 else "3:00 PM"
    if is_closer:
        return "11:00 PM" if day in {26, 27} else "9:30 PM"
    if start_minutes < 7 * 60:
        return "11:00 AM"
    if start_minutes < 10 * 60:
        return "2:00 PM"
    if start_minutes < 13 * 60:
        return "3:00 PM"
    return "11:00 PM" if day in {26, 27} else "9:30 PM"


def split_cell(value):
    text = str(value or "").strip()
    if not text or text.upper() in {"NA", "N/A"}:
        return []
    return [part.strip() for part in re.split(r"/", text) if part.strip()]


def parse_cell(value, default_role_name, date_key):
    text = str(value or "").strip()
    upper = text.upper().strip()
    if not text or upper in {"NA", "N/A"}:
        return []
    if upper == "RO":
        return [{"ro": True}]
    if "BUFFET" in upper and not re.search(r"\d", upper):
        return [{
            "roleName": "BOH Block",
            "start": "8:00 AM",
            "end": "4:00 PM",
            "untilVolume": False,
            "isCloser": False,
            "isFlexDouble": False,
            "isTraining": False,
            "notes": "Buffet",
        }]
    pieces = split_cell(text)
    parsed = []
    for index, piece in enumerate(pieces):
        piece_upper = piece.upper().replace(" ", "")
        if "BQT" in piece_upper and re.search(r"\d", piece_upper):
            normal_piece = re.sub(r"BQT|BANQUET", "", piece, flags=re.IGNORECASE)
            normal_shift = parse_time_token(normal_piece, date_key, default_role_name)
            if normal_shift:
                normal_shift["roleName"] = default_role_name
                normal_shift["notes"] = "; ".join(filter(None, [normal_shift.get("notes", ""), "Banquet after regular shift"]))
                parsed.append(normal_shift)
                parsed.append(banquet_after_shift(normal_shift, date_key))
            continue
        if piece_upper in {"BQT", "BANQUET"} and parsed:
            parsed[-1]["notes"] = "; ".join(filter(None, [parsed[-1].get("notes", ""), "Banquet after regular shift"]))
            parsed.append(banquet_after_shift(parsed[-1], date_key))
            continue
        if piece_upper in CROSS_REFERENCE_WORDS:
            continue
        parsed_piece = parse_time_token(piece, date_key, default_role_name)
        if parsed_piece:
            parsed.append(parsed_piece)
    return parsed


def banquet_after_shift(normal_shift, date_key):
    start = normal_shift["end"] if normal_shift.get("end") != "Until Volume" else normal_shift["start"]
    end = default_end_for("Banquet Server", start, date_key, False)
    if minutes(end) <= minutes(start):
        end = "11:59 PM"
    return {
        "roleName": "Banquet Server",
        "start": start,
        "end": end,
        "untilVolume": False,
        "isCloser": False,
        "isFlexDouble": False,
        "isTraining": False,
        "notes": "Banquet after regular shift",
    }


def iter_sections(workbook):
    sheet1 = workbook["Sheet1"]
    for row in range(3, 36):
        name = sheet1.cell(row, 1).value
        if name:
            for col, date_key in WEEK_DATES.items():
                yield name, "Server", date_key, sheet1.cell(row, col + 1).value
    sheet2 = workbook["Sheet2"]
    sections = [
        ("Host", 3, 12),
        ("Bartender", 17, 19),
        ("Busser", 23, 35),
    ]
    for role_name, start_row, end_row in sections:
        for row in range(start_row, end_row + 1):
            name = sheet2.cell(row, 1).value
            if name:
                for col, date_key in WEEK_DATES.items():
                    yield name, role_name, date_key, sheet2.cell(row, col + 1).value


def make_shift(data, employee, role, date_key, item):
    created = now_iso()
    notes = item.get("notes", "")
    return {
        "id": uid("shift"),
        "employeeId": employee["id"],
        "date": date_key,
        "shiftLabel": role["name"],
        "department": role["department"],
        "roleId": role["id"],
        "start": item["start"],
        "end": item["end"],
        "untilVolume": bool(item.get("untilVolume")),
        "meals": [],
        "notes": notes,
        "color": role.get("color", "#2563eb"),
        "createdAt": created,
        "updatedAt": created,
        "isCloser": bool(item.get("isCloser")),
        "isFlexDouble": bool(item.get("isFlexDouble")),
        "training": {
            "isTraining": bool(item.get("isTraining")),
            "traineeId": employee["id"] if item.get("isTraining") else "",
            "trainerId": "",
            "dayOverride": None,
        },
    }


def main():
    payload = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    data = payload["data"]
    import_created_employee_ids = {
        employee.get("id")
        for employee in data.get("employees", [])
        if employee.get("managerNotes") == "Created from old-school June 23-29 schedule import."
    }
    if import_created_employee_ids:
        data["employees"] = [
            employee for employee in data.get("employees", [])
            if employee.get("id") not in import_created_employee_ids
        ]
        data["shifts"] = [
            shift for shift in data.get("shifts", [])
            if shift.get("employeeId") not in import_created_employee_ids
        ]
        data["timeOffRequests"] = [
            request for request in data.get("timeOffRequests", [])
            if request.get("employeeId") not in import_created_employee_ids
        ]
    roles = role_by_name(data)
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    created_employees = []
    imported_shifts = []
    imported_ros = []
    skipped = []
    seen_shift_keys = set()
    seen_ro_keys = {
        (request.get("employeeId"), request.get("date"))
        for request in data.get("timeOffRequests", [])
        if WEEK_START <= request.get("date", "") <= WEEK_END
    }

    for source_name, default_role_name, date_key, cell_value in iter_sections(workbook):
        parsed_items = parse_cell(cell_value, default_role_name, date_key)
        if not parsed_items:
            continue
        employee = find_employee(data, str(source_name).strip(), created_employees)
        if not employee:
            skipped.append((str(source_name).strip(), date_key, cell_value, "No existing employee match"))
            continue
        for item in parsed_items:
            if item.get("ro"):
                key = (employee["id"], date_key)
                if key not in seen_ro_keys:
                    imported_ros.append({
                        "id": uid("timeoff"),
                        "employeeId": employee["id"],
                        "date": date_key,
                        "daypart": "All day",
                        "note": "RO from old-school schedule",
                        "source": "Old-school schedule import",
                        "createdAt": now_iso(),
                        "updatedAt": now_iso(),
                    })
                    seen_ro_keys.add(key)
                continue
            role = roles.get(item["roleName"].lower())
            if not role:
                skipped.append((source_name, date_key, cell_value, f"Missing role {item['roleName']}"))
                continue
            ensure_role_training(employee, role)
            key = (employee["id"], date_key, role["id"], item["start"])
            if key in seen_shift_keys:
                continue
            seen_shift_keys.add(key)
            imported_shifts.append(make_shift(data, employee, role, date_key, item))

    before_shifts = len(data.get("shifts", []))
    before_unassigned = len(data.get("unassignedShifts", []))
    data["shifts"] = [
        shift for shift in data.get("shifts", [])
        if not (WEEK_START <= shift.get("date", "") <= WEEK_END)
    ] + imported_shifts
    data["unassignedShifts"] = [
        shift for shift in data.get("unassignedShifts", [])
        if not (WEEK_START <= shift.get("date", "") <= WEEK_END)
    ]
    data["timeOffRequests"] = data.get("timeOffRequests", []) + imported_ros
    timestamp = now_iso()
    data["meta"]["updatedAt"] = timestamp
    payload["savedAt"] = timestamp
    payload["data"] = data

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f"restaurant-scheduler-data-before-june29-import-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"
    shutil.copy2(DATA_FILE, backup_path)
    DATA_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    print(json.dumps({
        "backup": str(backup_path),
        "removedAssignedShifts": before_shifts - len([s for s in data.get("shifts", []) if not (WEEK_START <= s.get("date", "") <= WEEK_END)]),
        "removedOpenShifts": before_unassigned - len(data.get("unassignedShifts", [])),
        "importedShifts": len(imported_shifts),
        "importedRequestOffs": len(imported_ros),
        "createdEmployees": created_employees,
        "skipped": skipped,
    }, indent=2))


if __name__ == "__main__":
    main()
